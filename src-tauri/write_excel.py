# -*- coding: utf-8 -*-
"""
Excel writer using openpyxl — supports BOTH modes:
  1. Full sheet write (legacy): {"sheets": [{"name": "...", "rows": [...]}]}
  2. Incremental changes:      {"sheets": [{"name": "...", "changes": [...], "insert_rows": [...]}]}
  3. Copy from source:         {"sourceFile": "...", "copyRows": [...]}
"""
import sys
import json

try:
    import openpyxl
except ImportError:
    print(
        "Missing Python dependency: openpyxl\n"
        "Install it with:\n"
        "  python -m pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)


def is_empty(v):
    return v is None or (isinstance(v, str) and v.strip() == "")


def set_cell(cell, val, formula=None):
    """Set cell value, with optional formula"""
    if formula:
        cell.value = formula
    elif is_empty(val):
        cell.value = None
    elif isinstance(val, (int, float)):
        cell.value = val
    elif isinstance(val, bool):
        cell.value = val
    else:
        cell.value = str(val)


def copy_cell_style(from_cell, to_cell):
    """Copy style from one cell to another"""
    if from_cell.has_style:
        to_cell.font = from_cell.font.copy()
        to_cell.border = from_cell.border.copy()
        to_cell.fill = from_cell.fill.copy()
        to_cell.number_format = from_cell.number_format
        to_cell.protection = from_cell.protection.copy()
        to_cell.alignment = from_cell.alignment.copy()


def detect_mode(data):
    """Detect if data is full-sheet, incremental, or copy-from-source"""
    sheets = data.get('sheets', [])
    if 'sourceFile' in data:
        return 'copy'
    if not sheets:
        return 'unknown'
    first = sheets[0]
    if 'changes' in first or 'insert_rows' in first:
        return 'incremental'
    if 'rows' in first:
        return 'full'
    return 'unknown'


def write_full(wb, data):
    """Full sheet write mode — writes all data to sheets"""
    for sheet_data in data['sheets']:
        sheet_name = sheet_data['name']
        rows = sheet_data['rows']

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)

        # Write all cells
        for ri, row in enumerate(rows, start=1):
            for ci, cell_data in enumerate(row, start=1):
                if isinstance(cell_data, dict):
                    val = cell_data.get('value')
                    formula = cell_data.get('formula')
                    set_cell(ws.cell(row=ri, column=ci), val, formula)
                else:
                    set_cell(ws.cell(row=ri, column=ci), cell_data)

        # Clear extra rows
        max_row_to_keep = len(rows)
        if ws.max_row > max_row_to_keep:
            for ri in range(max_row_to_keep + 1, ws.max_row + 1):
                for ci in range(1, ws.max_column + 1):
                    ws.cell(row=ri, column=ci).value = None

        # Delete extra columns (beyond what source data has)
        max_col_to_keep = max((len(row) for row in rows), default=0)
        if ws.max_column > max_col_to_keep:
            ws.delete_cols(max_col_to_keep + 1, ws.max_column - max_col_to_keep)


def write_incremental(wb, data):
    """Incremental mode — only modify changed cells"""
    for sheet_change in data['sheets']:
        sheet_name = sheet_change['name']
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

        # Apply cell changes
        for change in sheet_change.get('changes', []):
            val = change.get('value')
            formula = change.get('formula')
            set_cell(ws.cell(row=change['row'], column=change['col']), val, formula)

        # Insert new rows
        for row_data in sheet_change.get('insert_rows', []):
            new_row = ws.max_row + 1
            for ci, cell_data in enumerate(row_data, start=1):
                if isinstance(cell_data, dict):
                    val = cell_data.get('value')
                    formula = cell_data.get('formula')
                    set_cell(ws.cell(row=new_row, column=ci), val, formula)
                else:
                    set_cell(ws.cell(row=new_row, column=ci), cell_data)

        # Clear deleted rows
        for row_idx in sheet_change.get('delete_rows', []):
            for ci in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=ci).value = None


def write_copy_from_source(wb, data):
    """Copy mode - copy rows from source file while preserving formulas"""
    source_file = data['sourceFile']
    wb_source = openpyxl.load_workbook(source_file)

    for copy_op in data.get('copyRows', []):
        sheet_name = copy_op['sheetName']
        source_row = copy_op['sourceRow']
        target_row = copy_op['targetRow']

        if sheet_name not in wb_source.sheetnames:
            continue
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)

        ws_source = wb_source[sheet_name]
        ws_target = wb[sheet_name]

        # Ensure target row exists
        while target_row > ws_target.max_row:
            ws_target.append([])

        # Copy all cells from source to target
        if source_row <= ws_source.max_row:
            for col_idx in range(1, ws_source.max_column + 1):
                source_cell = ws_source.cell(row=source_row, column=col_idx)
                target_cell = ws_target.cell(row=target_row, column=col_idx)
                target_cell.value = source_cell.value
                copy_cell_style(source_cell, target_cell)

    wb_source.close()


def main():
    if len(sys.argv) < 3:
        print("Usage: python write_excel.py <file_path> <json_file>", file=sys.stderr)
        sys.exit(1)

    file_path = sys.argv[1]
    json_path = sys.argv[2]

    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    keep_vba = file_path.lower().endswith('.xlsm')
    wb = openpyxl.load_workbook(file_path, keep_vba=keep_vba)

    mode = detect_mode(data)
    if mode == 'full':
        write_full(wb, data)
    elif mode == 'incremental':
        write_incremental(wb, data)
    elif mode == 'copy':
        write_copy_from_source(wb, data)
    else:
        print("Unknown data format", file=sys.stderr)
        sys.exit(1)

    wb.save(file_path)
    print("OK", flush=True)


if __name__ == '__main__':
    main()
