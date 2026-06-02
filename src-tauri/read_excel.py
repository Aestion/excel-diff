"""
Read Excel file - DEAD SIMPLE VERSION
"""
import sys
import json
import openpyxl


def get_val(cell):
    """Get cell value, no matter what"""
    if cell.value is None:
        return None
    if isinstance(cell.value, bool):
        return cell.value
    if isinstance(cell.value, int):
        return cell.value
    if isinstance(cell.value, float):
        if cell.value.is_integer():
            return int(cell.value)
        return cell.value
    return str(cell.value)


def main():
    if len(sys.argv) < 2:
        print("Usage: python read_excel.py <file_path>", file=sys.stderr)
        sys.exit(1)

    file_path = sys.argv[1]

    try:
        # Open file - NO data_only! Just read whatever is there
        wb = openpyxl.load_workbook(file_path, read_only=False, data_only=False, keep_vba=False)
        sheet_names = wb.sheetnames
        sheets = []

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            max_row = ws.max_row
            max_col = ws.max_column

            # Read headers
            columns = []
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col_idx)
                val = get_val(cell)
                name = val or f"Column {col_idx}"
                columns.append({"index": col_idx - 1, "name": str(name), "dataType": "mixed"})

            # Read rows
            rows = []
            for row_idx in range(1, max_row + 1):
                row_data = []
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = get_val(cell)
                    cell_obj = {"value": val}

                    # If it's a formula, store formula separately.
                    # Keep the formula string as the display value initially.
                    # The second pass (data_only=True) will try to replace it
                    # with the cached computed value if available.
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        cell_obj["formula"] = cell.value

                    row_data.append(cell_obj)
                rows.append(row_data)

            sheets.append({"name": sheet_name, "columns": columns, "rows": rows})

        wb.close()

        # Now try a second pass with data_only=True to get missing values
        try:
            wb2 = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
            for sheet_idx, sheet_name in enumerate(sheet_names):
                if sheet_name in wb2.sheetnames:
                    ws2 = wb2[sheet_name]
                    sheet_data = sheets[sheet_idx]
                    for row_idx in range(len(sheet_data["rows"])):
                        row_data = sheet_data["rows"][row_idx]
                        for col_idx in range(len(row_data)):
                            if "formula" in row_data[col_idx]:
                                cell2 = ws2.cell(row=row_idx + 1, column=col_idx + 1)
                                val2 = get_val(cell2)
                                if val2 is not None:
                                    row_data[col_idx]["value"] = val2
            wb2.close()
        except:
            pass

        result = {
            "filePath": file_path,
            "sheets": sheets,
            "sheetNames": sheet_names
        }

        json_str = json.dumps(result, ensure_ascii=True)
        sys.stdout.buffer.write(json_str.encode('utf-8'))
        sys.stdout.flush()

    except Exception as e:
        print(f"Error: {str(e)}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
